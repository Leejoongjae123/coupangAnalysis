import datetime
import json

import pandas as pd
import requests
import requests
from bs4 import BeautifulSoup
import time
import random
import pprint
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import openpyxl
import numpy as np
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import getmac
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from window import Ui_MainWindow
import sys
import urllib.parse

def GetReview(reviewCount,productId):
    count = 1
    reviewList=[]
    endFlag=False
    while True:


        if endFlag==True:
            break
        
        for page in range(1,100):
            cookies = {
                'PCID': '56954098591750730901126',
                'x-coupang-accept-language': 'ko-KR',
                '_fbp': 'fb.1.1698752837926.1707360410',
                'gd1': 'Y',
                'MARKETID': '56954098591750730901126',
                'x-coupang-target-market': 'KR',
                'sid': '14c2ba9ac1ae4f05bbaa1288b73e479c24c52d68',
                'searchKeyword': '%ED%81%AC%EB%A6%BC',
                'searchKeywordType': '%7B%22%ED%81%AC%EB%A6%BC%22%3A0%7D',
                'overrideAbTestGroup': '%5B%5D',
                'bm_sz': '0B8ECB1E586103F15D5578DB82FCA948~YAAQPw3VF5xvVFaMAQAApPMqjRY8/XB+p1liFY/UnvY260ScoIliSdJJ/m8wgd/iVf2rcxKkLFosW+lHQKEeassb0+2J6wOZASQNPdR52RFToRiaai3/msolQaVEDPiRK9iiikj71beFr8G18Wm6p2Fb99hVyZ+ayZdpSNEvTs3nvdSdWy0TozhV7M/hbre1mtU69b+Qn4Z90XxafiPfNYSHoWjaesuMJ7AHOfkcd1pP21qxl/ZXKPdaNvGrgEv5tHjW7eAQZwKqlGEC7n1f2lN7zOA2iA/VSdeFv4WYQBkdUI19HH47tFM1QFFyGFxyCpKVoSuJiFutqYigaw==~4276530~3618883',
                # 'ak_bmsc': '2576F0C14DA61F936ADA1A89BAF7ADAC~000000000000000000000000000000~YAAQPw3VF9lwVFaMAQAAt/cqjRZOL+c9CtM8b1vNZwu0of2VoJ1k+SUKn4r/1v1gRqkS2jQWfsLCFA9fJqgBwKrC0Gr0MfmtVbEngcbDs9616BUg9usrsGNnfqZRXyB4zI8FKhiSBRr07S/q15JddvoTd1AHUH1RS0BpDGa9wBvb++pK7VHAF5V0T/s9VXF0J9+mLDudNMY6les0ptaotw4JJ5exsgJfAoYaIw/yh+9KOkNtPJXxFDMZRdCMebFYhoIp16/Y6C3HMHIoispM2Agm0jWSasaduoAhohASEzWxJEmQnlax98pYxRgElEsSflkw0OGgFgnVA5X8X+qvFnRFrI1YiRhXB4ZXiapU4JqWLyZOGlEuYUdVNLREu0ZEZPzWtsGa+5DKYCh3fFnyLCMV0Xm5JfdCrSbCDQX/4ePM9Hm6eOuvwLco4FXNRPQgARehVThIK90PK+4DQK9EP2NsMfWlt8y33ksHPQ2snLwWrDbbSYVyMrBdpGo0CQA=',
                '_abck': '6E1F205650720FF98BACA769F458FD6A~0~YAAQPw3VF552VFaMAQAA8gYrjQuaKaZaxOhlWvaL1znDyBzYqMdIW5trcUsNdDmTsaiuUVG3+gh71LUQGFl++idVZXSw4DoQCtWW/srLolu9n7Klk3Imx18tYr1oxFF+WmtV0hxu2Ww8R5pvsDe/989gUCQKVPUm5WT4amjQWtXte3scqb7jAzDPJRtyNu6BA8Go4vDoHmzGzfA+4yBNh0G72HSuNos3iyFvyK4LpR9JZc/zirfib9TO+QBumjtZT1wn8KpVyZy6RLftf4b2selA683zWd2DVIUeor7ual6cFD5U/qg6eWpJfABK8I8ZeQCGH5US4GFDyznlFAXKzIdWj3NDDBBZczEx7W7JGSDHCwc70+1aTpgvZYTU5wFVqvFwLArZbcu7YgAzn8/Q4hkpts6B9/U5f8LqZfDQpQVuc0GUkSe+~-1~-1~-1',
                'cto_bundle': 'E_XeDV9rUXJ2TDF4NWRLeGxud0pJRmg1Y3NEOEhKMXlPVlFSOU5EbmJ6ZGpuR3JLdFA0ZDBTTVhrRWsxSWY5eUl2dGVacEdFSEZ1ZHdOZFNDVzBwUGlYZEVqeU11cDdWT3VrSGdzMzFTbjNVbjVnQUFoWkhKekdyeGtKVGp0Umt4cFNJMFp1bUElMkZ3dmh4eDQ5RUhBYW16U3BVQzRnQk9YYmclMkZlZWhKd1BBY1BPYlc3bzlxJTJGNjFFRlFPVmtDbWg5TDFSbGxYaEFFb3loZXRQekwlMkJEcGVpY2JrQWs4cVlYNW41Nm5rRXRIaXM1SDRYNllISEJBbG55bGFIY2NQeXJORGwyV0g',
                'baby-isWide': 'wide',
                'bm_sv': 'FA56C55F114E6465FFB5E0B759060363~YAAQPw3VFy3MVFaMAQAAtN8rjRbtZ9cSlWe6JpMsMMHr6Xfz7ZEIWX+pU1ey0IGsmgAwlk0+gGOcmDfenykk/SHelQWN3Sf6A66DiZGGQedH/WAhC66K+7lu62T34n1GIgpo2lzVo5YfVRWbfqIjTXWOq5tGs4zRRQwZQ3Bv7KJ9/Qf3Vp9i4uQZjWvzcZAAsArCHtizaUWAZOyPxbRDiFy4LLAo225ABixp70lco53k2JXUbgo/HBGE5PuR2iIryPE=~1',
            }

            headers = {
                'authority': 'www.coupang.com',
                'accept': '*/*',
                'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                # 'cookie': 'PCID=56954098591750730901126; x-coupang-accept-language=ko-KR; _fbp=fb.1.1698752837926.1707360410; gd1=Y; MARKETID=56954098591750730901126; x-coupang-target-market=KR; sid=14c2ba9ac1ae4f05bbaa1288b73e479c24c52d68; searchKeyword=%ED%81%AC%EB%A6%BC; searchKeywordType=%7B%22%ED%81%AC%EB%A6%BC%22%3A0%7D; overrideAbTestGroup=%5B%5D; bm_sz=0B8ECB1E586103F15D5578DB82FCA948~YAAQPw3VF5xvVFaMAQAApPMqjRY8/XB+p1liFY/UnvY260ScoIliSdJJ/m8wgd/iVf2rcxKkLFosW+lHQKEeassb0+2J6wOZASQNPdR52RFToRiaai3/msolQaVEDPiRK9iiikj71beFr8G18Wm6p2Fb99hVyZ+ayZdpSNEvTs3nvdSdWy0TozhV7M/hbre1mtU69b+Qn4Z90XxafiPfNYSHoWjaesuMJ7AHOfkcd1pP21qxl/ZXKPdaNvGrgEv5tHjW7eAQZwKqlGEC7n1f2lN7zOA2iA/VSdeFv4WYQBkdUI19HH47tFM1QFFyGFxyCpKVoSuJiFutqYigaw==~4276530~3618883; ak_bmsc=2576F0C14DA61F936ADA1A89BAF7ADAC~000000000000000000000000000000~YAAQPw3VF9lwVFaMAQAAt/cqjRZOL+c9CtM8b1vNZwu0of2VoJ1k+SUKn4r/1v1gRqkS2jQWfsLCFA9fJqgBwKrC0Gr0MfmtVbEngcbDs9616BUg9usrsGNnfqZRXyB4zI8FKhiSBRr07S/q15JddvoTd1AHUH1RS0BpDGa9wBvb++pK7VHAF5V0T/s9VXF0J9+mLDudNMY6les0ptaotw4JJ5exsgJfAoYaIw/yh+9KOkNtPJXxFDMZRdCMebFYhoIp16/Y6C3HMHIoispM2Agm0jWSasaduoAhohASEzWxJEmQnlax98pYxRgElEsSflkw0OGgFgnVA5X8X+qvFnRFrI1YiRhXB4ZXiapU4JqWLyZOGlEuYUdVNLREu0ZEZPzWtsGa+5DKYCh3fFnyLCMV0Xm5JfdCrSbCDQX/4ePM9Hm6eOuvwLco4FXNRPQgARehVThIK90PK+4DQK9EP2NsMfWlt8y33ksHPQ2snLwWrDbbSYVyMrBdpGo0CQA=; _abck=6E1F205650720FF98BACA769F458FD6A~0~YAAQPw3VF552VFaMAQAA8gYrjQuaKaZaxOhlWvaL1znDyBzYqMdIW5trcUsNdDmTsaiuUVG3+gh71LUQGFl++idVZXSw4DoQCtWW/srLolu9n7Klk3Imx18tYr1oxFF+WmtV0hxu2Ww8R5pvsDe/989gUCQKVPUm5WT4amjQWtXte3scqb7jAzDPJRtyNu6BA8Go4vDoHmzGzfA+4yBNh0G72HSuNos3iyFvyK4LpR9JZc/zirfib9TO+QBumjtZT1wn8KpVyZy6RLftf4b2selA683zWd2DVIUeor7ual6cFD5U/qg6eWpJfABK8I8ZeQCGH5US4GFDyznlFAXKzIdWj3NDDBBZczEx7W7JGSDHCwc70+1aTpgvZYTU5wFVqvFwLArZbcu7YgAzn8/Q4hkpts6B9/U5f8LqZfDQpQVuc0GUkSe+~-1~-1~-1; cto_bundle=E_XeDV9rUXJ2TDF4NWRLeGxud0pJRmg1Y3NEOEhKMXlPVlFSOU5EbmJ6ZGpuR3JLdFA0ZDBTTVhrRWsxSWY5eUl2dGVacEdFSEZ1ZHdOZFNDVzBwUGlYZEVqeU11cDdWT3VrSGdzMzFTbjNVbjVnQUFoWkhKekdyeGtKVGp0Umt4cFNJMFp1bUElMkZ3dmh4eDQ5RUhBYW16U3BVQzRnQk9YYmclMkZlZWhKd1BBY1BPYlc3bzlxJTJGNjFFRlFPVmtDbWg5TDFSbGxYaEFFb3loZXRQekwlMkJEcGVpY2JrQWs4cVlYNW41Nm5rRXRIaXM1SDRYNllISEJBbG55bGFIY2NQeXJORGwyV0g; baby-isWide=wide; bm_sv=FA56C55F114E6465FFB5E0B759060363~YAAQPw3VFy3MVFaMAQAAtN8rjRbtZ9cSlWe6JpMsMMHr6Xfz7ZEIWX+pU1ey0IGsmgAwlk0+gGOcmDfenykk/SHelQWN3Sf6A66DiZGGQedH/WAhC66K+7lu62T34n1GIgpo2lzVo5YfVRWbfqIjTXWOq5tGs4zRRQwZQ3Bv7KJ9/Qf3Vp9i4uQZjWvzcZAAsArCHtizaUWAZOyPxbRDiFy4LLAo225ABixp70lco53k2JXUbgo/HBGE5PuR2iIryPE=~1',
                'referer': 'https://www.coupang.com/vp/products/1901239628?vendorItemId=71216468367&sourceType=HOME_RELATED_ADS&searchId=feed-8240d194ecdb4ce8b47b6b1bc442eb35-related_ads&clickEventId=71fdc910-a01c-11ee-ba73-a012bbd22ed1&isAddedCart=',
                'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'x-requested-with': '',
            }

            params = {
                'productId': productId,
                'page': page,
                'size': '10',
                'sortBy': 'DATE_DESC',
                'ratings': '',
                'q': '',
                # 'viRoleCode': '2',
                'ratingSummary': 'true',
            }

            response = requests.get('https://www.coupang.com/vp/product/reviews', params=params, cookies=cookies, headers=headers)
            # print(response.text)
            print("status_code:",response.status_code)
            soup=BeautifulSoup(response.text,'lxml')

            regiDates=soup.find_all("div",attrs={'class':'sdp-review__article__list__info__product-info__reg-date'})
            if len(regiDates)==0:
                print("리뷰없음")
                endFlag=True
                break
            for regiDate in regiDates:
                regiDateText=regiDate.get_text()
                print("regiDateText:",regiDateText,"/ regiDateText_TYPE:",type(regiDateText))
                reviewList.append(regiDateText)



            print("page:",page,"/ page_TYPE:",type(page))
            if len(reviewList)>=reviewCount:
                reviewList=reviewList[:reviewCount]
                endFlag=True
                break



            count+=1
            time.sleep(random.randint(10,20)*0.1)
            if len(regiDates)<10:
                print("리뷰끝남")
                endFlag=True
                break

        with open('source/reviewList.json', 'w',encoding='utf-8-sig') as f:
            json.dump(reviewList, f, indent=2,ensure_ascii=False)

        try:
            firstReview=reviewList[-1]
        except:
            firstReview=""
        print("firstReview:",firstReview,"/ firstReview_TYPE:",type(firstReview))
    # 날짜 범위 설정 (2023년 12월 20일부터 2023년 12월 27일까지)
    daysAgo=7
    start_date = datetime.datetime.now()-datetime.timedelta(days=daysAgo+1)
    end_date = datetime.datetime.now()
    # print("start_date:",start_date,"/ start_date_TYPE:",type(start_date))
    # print("end_date:",end_date,"/ end_date_TYPE:",type(end_date))
    # 날짜 범위에 속하는 날짜의 갯수를 카운트
    count7 = sum(start_date <= datetime.datetime.strptime(date, "%Y.%m.%d") <= end_date for date in reviewList)
    print("count7:",count7,"/ count_TYPE:",type(count7))

    daysAgo=28
    start_date = datetime.datetime.now()-datetime.timedelta(days=daysAgo+1)
    end_date = datetime.datetime.now()
    # print("start_date:",start_date,"/ start_date_TYPE:",type(start_date))
    # print("end_date:",end_date,"/ end_date_TYPE:",type(end_date))
    # 날짜 범위에 속하는 날짜의 갯수를 카운트
    count28 = sum(start_date <= datetime.datetime.strptime(date, "%Y.%m.%d") <= end_date for date in reviewList)
    print("count28:",count28,"/ count_TYPE:",type(count28))


    return reviewList,count7,count28,firstReview

def create_url(baseurl, params):
    # params 사전을 쿼리 문자열로 변환
    query_string = urllib.parse.urlencode(params)
    # 기본 URL과 쿼리 문자열을 결합
    url = f"{baseurl}?{query_string}"
    return url

def GetSearch(input):
    cookies = {
        'PCID': '56954098591750730901126',
        'x-coupang-accept-language': 'ko-KR',
        '_fbp': 'fb.1.1698752837926.1707360410',
        'gd1': 'Y',
        'MARKETID': '56954098591750730901126',
        'x-coupang-target-market': 'KR',
        'sid': '14c2ba9ac1ae4f05bbaa1288b73e479c24c52d68',
        'overrideAbTestGroup': '%5B%5D',
        'bm_sz': '79E7DE317CF51BEC3D34E934D0EF084F~YAAQZGHKFwrZzW+MAQAAVLCykBboCf4pm5AwQ3H1DBUY8fh7wwqWb14lpUdWjF1HS37pcw5aM6AmsaNan+Ea07ux4+KhtDUvfITER4r30OCMbjeGwZmy3NC6rKdr89e8ngtY36kjtD1zDqgXVbOyB2Uqqtbmtm8FeqoXXPYlOKEyMvHevxhcg2kXqCZhKMr85OePkEJj/9O2wkcNOebnBx/1/m4TrtpGv6vHCwtaQRJ5sA9QzhrCTgJoelWGyv7eG4j4fFCp7I3SwCNfxCG1EHEmHKNClsdu0mjEWzlZAM5BoR5rY31xDt8XJkVYyAXqoaMZE4xWSdYxzrE30Q==~3294790~3687732',
        # 'ak_bmsc': '56F803B46AC96D2B3E7FE0636DC50632~000000000000000000000000000000~YAAQZGHKFwHbzW+MAQAAarSykBbqtdcxwMantf4fcQ3BoisRsfWUuVqbKV9AdYedi/B5AV/99z92/N4sfLLH99Vr6jRdlRqWhtw/HC86HTwHDNRP1R2f61UFxflauanwlvB5Vlajg+yiD55Xkn6DwUNhXPaCxOYe00xDLqeM5JSr6lkLMhJbbNbhycD9KheJdjP2hfeKJA6h10LF19zAcjxN3zuA6U27TPCOIsNF0qLLTg6QfB7TT5m3MpqqTKP71qsqn84GgrP8M7TELlQidI3vGfpl/Hxy3rUKf8l2WIitRQemkVc+O7ZhgQwXyLbqHHe2BNCLZpNq/ub7U9JB2+IRIjNXnFonh+rPi0QI0U6FYe/b5sLapzv3tp2cRGXLl7zxibsief4YEL7kZhgQ6S6vskn06d04KoQjTe8nQcaCgKBwaF6KoPBBOyZA4VJXtf//DV9XwuipJtTSwOppPKEeQCbUsE6L6Oh+NYu2zcZ2o8TRvDfv4ub8Gax6j80=',
        'searchKeyword': '%ED%81%AC%EB%A6%BC%7C%EC%88%98%EB%B6%84%ED%81%AC%EB%A6%BC',
        'searchKeywordType': '%7B%22%ED%81%AC%EB%A6%BC%22%3A0%7D%7C%7B%22%EC%88%98%EB%B6%84%ED%81%AC%EB%A6%BC%22%3A0%7D',
        '_abck': '6E1F205650720FF98BACA769F458FD6A~0~YAAQZGHKF4vkzW+MAQAArcqykAvj9hfqb3uyMGcBr8vntLWsmV9ZIT069tphxC0Aib2hwUwj0HTUHKtF9fxvHFlK0XwZWog7K1rvDyUWhv+sXHN5jAWGjH8ei6Sr+iVZVYvT5icxFtKiV6vPafZIx47WxLSAg0K0sl3BIehY3EtVx39CxpVM4R5FroqKaEdpqxh/VwRHrfxRB6Xz+tktKYiVa9Jcv3TG27Jzybny33Pe6yzXQSDZ/Ui7vYx6wNNJIg9BIZT3rzlfg7tdFJSBBJP1XW+hKCPcalFKU+A9mb6PdhC7TooM3P+evfaPgu7x+O5YnDJ7Bz65selEPBjnb6FubRAyAPmyDqj9AXMcuLg2W4e7z29E0ExNsqDfX5BqHjcAY54LzYMz996meOzOjALdTNAWGNoN5DzDUiNSDOgwL9lQDM6o~-1~-1~-1',
        'baby-isWide': 'small',
        'cto_bundle': 'Jd6AZ19rUXJ2TDF4NWRLeGxud0pJRmg1Y3NINjlpOERLa1F2RmlVNURRajVPUHlFbkliOU92ekFNSnNuWXFGbXVFZW54OEJVMVM5WlJaTkdmMGpXRXglMkZCajNISk1zV1ZYb25VOEM5aTdjUEw2VHVsbmE3WnY3d083bFhrWmFReFBPVURNSldNeG9qVCUyQnk4b2VWa2ppbk5RZ01kamdrZGFDMmsydTZiYzJEQ2tsbTVrSVhHbVNNS1JGJTJGSkxiZlJFYW9pTmpZOEc1Y3U5SlBDNkxQc2J4YU1QZHNMUiUyQlNmdjFodzJqNnJXTmN2ZDF2Z2phdng1YWZTT3NmZVVkQWh2ZDU5anM',
        'bm_sv': '304BEF3CF23C8106790815B2D1276599~YAAQR2HKFwA5yWaMAQAA4/60kBbJEbfsDTZh7io8bfkiAQ+50J0m2FlzLY59XsOJnnXVP1R+/91XeGJlibwQ0OiJKOEy20fs/NCWD8JtRSE0Ysyy/9T0uFGQnZeKoewZ/HUOMNqFUwtpsWkE/XY88RDk6q7VB6abBiBc/oj12RjI/+v8DhGnP6o9OJWDTZhQ7JwfoW6covCAaMF7LjpMTAzUSET32czi0Wf9xMG6ltf49MVcFx1kcgjodh9uqQoToFQ=~1',
    }

    headers = {
        'authority': 'www.coupang.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        # 'cookie': 'PCID=56954098591750730901126; x-coupang-accept-language=ko-KR; _fbp=fb.1.1698752837926.1707360410; gd1=Y; MARKETID=56954098591750730901126; x-coupang-target-market=KR; sid=14c2ba9ac1ae4f05bbaa1288b73e479c24c52d68; overrideAbTestGroup=%5B%5D; bm_sz=79E7DE317CF51BEC3D34E934D0EF084F~YAAQZGHKFwrZzW+MAQAAVLCykBboCf4pm5AwQ3H1DBUY8fh7wwqWb14lpUdWjF1HS37pcw5aM6AmsaNan+Ea07ux4+KhtDUvfITER4r30OCMbjeGwZmy3NC6rKdr89e8ngtY36kjtD1zDqgXVbOyB2Uqqtbmtm8FeqoXXPYlOKEyMvHevxhcg2kXqCZhKMr85OePkEJj/9O2wkcNOebnBx/1/m4TrtpGv6vHCwtaQRJ5sA9QzhrCTgJoelWGyv7eG4j4fFCp7I3SwCNfxCG1EHEmHKNClsdu0mjEWzlZAM5BoR5rY31xDt8XJkVYyAXqoaMZE4xWSdYxzrE30Q==~3294790~3687732; ak_bmsc=56F803B46AC96D2B3E7FE0636DC50632~000000000000000000000000000000~YAAQZGHKFwHbzW+MAQAAarSykBbqtdcxwMantf4fcQ3BoisRsfWUuVqbKV9AdYedi/B5AV/99z92/N4sfLLH99Vr6jRdlRqWhtw/HC86HTwHDNRP1R2f61UFxflauanwlvB5Vlajg+yiD55Xkn6DwUNhXPaCxOYe00xDLqeM5JSr6lkLMhJbbNbhycD9KheJdjP2hfeKJA6h10LF19zAcjxN3zuA6U27TPCOIsNF0qLLTg6QfB7TT5m3MpqqTKP71qsqn84GgrP8M7TELlQidI3vGfpl/Hxy3rUKf8l2WIitRQemkVc+O7ZhgQwXyLbqHHe2BNCLZpNq/ub7U9JB2+IRIjNXnFonh+rPi0QI0U6FYe/b5sLapzv3tp2cRGXLl7zxibsief4YEL7kZhgQ6S6vskn06d04KoQjTe8nQcaCgKBwaF6KoPBBOyZA4VJXtf//DV9XwuipJtTSwOppPKEeQCbUsE6L6Oh+NYu2zcZ2o8TRvDfv4ub8Gax6j80=; searchKeyword=%ED%81%AC%EB%A6%BC%7C%EC%88%98%EB%B6%84%ED%81%AC%EB%A6%BC; searchKeywordType=%7B%22%ED%81%AC%EB%A6%BC%22%3A0%7D%7C%7B%22%EC%88%98%EB%B6%84%ED%81%AC%EB%A6%BC%22%3A0%7D; _abck=6E1F205650720FF98BACA769F458FD6A~0~YAAQZGHKF4vkzW+MAQAArcqykAvj9hfqb3uyMGcBr8vntLWsmV9ZIT069tphxC0Aib2hwUwj0HTUHKtF9fxvHFlK0XwZWog7K1rvDyUWhv+sXHN5jAWGjH8ei6Sr+iVZVYvT5icxFtKiV6vPafZIx47WxLSAg0K0sl3BIehY3EtVx39CxpVM4R5FroqKaEdpqxh/VwRHrfxRB6Xz+tktKYiVa9Jcv3TG27Jzybny33Pe6yzXQSDZ/Ui7vYx6wNNJIg9BIZT3rzlfg7tdFJSBBJP1XW+hKCPcalFKU+A9mb6PdhC7TooM3P+evfaPgu7x+O5YnDJ7Bz65selEPBjnb6FubRAyAPmyDqj9AXMcuLg2W4e7z29E0ExNsqDfX5BqHjcAY54LzYMz996meOzOjALdTNAWGNoN5DzDUiNSDOgwL9lQDM6o~-1~-1~-1; baby-isWide=small; cto_bundle=Jd6AZ19rUXJ2TDF4NWRLeGxud0pJRmg1Y3NINjlpOERLa1F2RmlVNURRajVPUHlFbkliOU92ekFNSnNuWXFGbXVFZW54OEJVMVM5WlJaTkdmMGpXRXglMkZCajNISk1zV1ZYb25VOEM5aTdjUEw2VHVsbmE3WnY3d083bFhrWmFReFBPVURNSldNeG9qVCUyQnk4b2VWa2ppbk5RZ01kamdrZGFDMmsydTZiYzJEQ2tsbTVrSVhHbVNNS1JGJTJGSkxiZlJFYW9pTmpZOEc1Y3U5SlBDNkxQc2J4YU1QZHNMUiUyQlNmdjFodzJqNnJXTmN2ZDF2Z2phdng1YWZTT3NmZVVkQWh2ZDU5anM; bm_sv=304BEF3CF23C8106790815B2D1276599~YAAQR2HKFwA5yWaMAQAA4/60kBbJEbfsDTZh7io8bfkiAQ+50J0m2FlzLY59XsOJnnXVP1R+/91XeGJlibwQ0OiJKOEy20fs/NCWD8JtRSE0Ysyy/9T0uFGQnZeKoewZ/HUOMNqFUwtpsWkE/XY88RDk6q7VB6abBiBc/oj12RjI/+v8DhGnP6o9OJWDTZhQ7JwfoW6covCAaMF7LjpMTAzUSET32czi0Wf9xMG6ltf49MVcFx1kcgjodh9uqQoToFQ=~1',
        'referer': 'https://www.coupang.com/',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }

    params = {
        'rocketAll': 'false',
        'searchId': 'c587db561ab144659dee73c8333e9137',
        'q': input,
        'brand': '',
        'offerCondition': '',
        'filter': '',
        'availableDeliveryFilter': '',
        'filterType': '',
        'isPriceRange': 'false',
        'priceRange': '',
        'minPrice': '',
        'maxPrice': '',
        'page': '1',
        'trcid': '',
        'traid': '',
        'filterSetByUser': 'true',
        'channel': 'user',
        'backgroundColor': '',
        'searchProductCount': '1248109',
        'component': '',
        'rating': '0',
        'sorter': 'scoreDesc',
        'listSize': '72',
    }
    baseUrl='https://www.coupang.com/np/search'
    response = requests.get(baseUrl, params=params, cookies=cookies, headers=headers)
    soup=BeautifulSoup(response.text,'lxml')

    searchUrl=create_url(baseUrl,params)

    # BeautifulSoup 객체 생성
    soup = BeautifulSoup(response.text, 'lxml')
    print(soup.prettify())
    # id가 'productList'인 ul 태그 찾기
    product_group = soup.find("ul", attrs={'id': 'productList'})

    # 해당 ul 태그 내부에서 클래스가 'search-product best-seller-carousel-item'인 모든 li 태그 찾기 및 제거
    for li in product_group.find_all("li", class_="search-product best-seller-carousel-item"):
        li.decompose()
    for li in product_group.find_all("li", class_="sdw-aging"):
        li.decompose()



    items=soup.find("ul", attrs={'id': 'productList'}).find_all('li')

    count=1

    rocketNfresh=0
    sellerrocket=0
    importrocket=0
    wing=0
    ad=0
    normal=0

    rocketNfreshReview=0
    sellerrocketReview = 0
    importrocketReview = 0
    wingReview = 0
    adReview = 0
    normalReview = 0
    priceList=[]
    productNoList=[]
    productInfoList = []
    for index,item in enumerate(items):
        try:
            title=item.find("div",attrs={'class':'name'}).get_text().replace("\n","").strip()
            text="{}번째 : {}".format(count,title)
            # print(text)
        except:
            print("없음")
            continue
        try:
            price=int(item.find('strong',attrs={'class':'price-value'}).get_text().replace(",",""))
        except:
            price=0
        try:
            url='https://www.coupang.com'+item.find('a',attrs={'class':'search-product-link'})['href']
        except:
            url=""
        print("url:",url)
        # print("price:",price)
        priceList.append(price)
        isAdLength=len(item.find_all('span',attrs={'class':'ad-badge-text'}))
        isAd=False
        isAdLength = len(item.find_all('span', attrs={'class': 'ad-badge-text'}))
        isAd = False

        try:
            productNo=item.find('a',attrs={'class':'search-product-link'})['href'].split("/")[3].split("?")[0]

        except:
            productNo=""
        productNoList.append(productNo)
        print("productNo:",productNo,"/ productNo_TYPE:",type(productNo))

        try:
            productName=item.find("div",attrs={'class':'name'}).get_text().strip()
        except:
            productName=""

        try:
            arrivalInfo=item.find("span",attrs={'class':'arrival-info'}).find_all('em')[0].get_text().strip()
        except:
            arrivalInfo=""
        print("arrivalInfo:",arrivalInfo,"/ arrivalInfo_TYPE:",type(arrivalInfo))

        if isAdLength >= 1:
            ad += 1
            print("광고라건너뜀")
            print("ad:",ad,"/ ad_TYPE:",type(ad))
            continue
        else:
            priceList.append(price)

        category="일반"
        try:
            badge=item.find('span',attrs={'class':'badge rocket'}).find('img')['src']
            if badge.find("fresh")>=0:
                rocketNfresh+=1
                print("후레시{}".format(rocketNfresh))
                try:
                    rating=int(item.find('span',attrs={'class':'rating-total-count'}).get_text().replace("(","").replace(")",""))
                except:
                    rating=0
                print("rating:",rating)
                rocketNfreshReview+=rating
                category="로켓후레쉬"
            elif badge.find("rocket_large")>=0:
                rocketNfresh+=1
                print("로켓{}".format(rocketNfresh))
                try:
                    rating=int(item.find('span',attrs={'class':'rating-total-count'}).get_text().replace("(","").replace(")",""))
                except:
                    rating=0
                print("rating:",rating)
                rocketNfreshReview += rating
                category="로켓배송"
            elif badge.find("Merchant")>=0:
                sellerrocket+=1
                print("판매자{}".format(sellerrocket))
                try:
                    rating=int(item.find('span',attrs={'class':'rating-total-count'}).get_text().replace("(","").replace(")",""))
                except:
                    rating=0
                print("rating:",rating)
                sellerrocketReview+=rating
                category="판매자로켓"
            elif badge.find("global")>=0:
                importrocket+=1
                print("직구{}".format(importrocket))
                try:
                    rating=int(item.find('span',attrs={'class':'rating-total-count'}).get_text().replace("(","").replace(")",""))
                except:
                    rating=0
                print("rating:",rating)
                importrocket+=rating
                category="로켓직구"



        except:
            normal+=1
            print("일반{}".format(normal))
            category = "일반상품"
            try:
                rating = int(item.find('span', attrs={'class': 'rating-total-count'}).get_text().replace("(", "").replace(")",
                                                                                                                 ""))
            except:
                rating = 0
            print("rating:", rating)
            normalReview+=rating

        productInfo = {'rank': count, 'keyword': input, 'productName': title, 'price': price,'category':category,'arrivalInfo': arrivalInfo,'productNo':productNo,'isAd':isAdLength,'totalReview':rating,'url':url}
        print("productInfo:",productInfo,"/ productInfo_TYPE:",type(productInfo))
        productInfoList.append(productInfo)
        count += 1


        print("==============")
    print("normal:",normal,"/ normal_TYPE:",type(normal))
    print("rocketNfresh:",rocketNfresh,"/ rocketNfresh_TYPE:",type(rocketNfresh))
    print("sellerrocket:",sellerrocket,"/ sellerrocket_TYPE:",type(sellerrocket))
    print("importrocket:",importrocket,"/ importrocket_TYPE:",type(importrocket))
    print("wing:",wing,"/ wing_TYPE:",type(wing))
    print("ad:",ad,"/ ad_TYPE:",type(ad))
    maxPrice=max(priceList)
    print("maxPrice:",maxPrice,"/ maxPrice_TYPE:",type(maxPrice))
    minPrice=min(priceList)
    print("minPrice:",minPrice,"/ minPrice_TYPE:",type(minPrice))
    diffPrice=maxPrice-minPrice
    print("diffPrice:",diffPrice,"/ diffPrice_TYPE:",type(diffPrice))
    totalReview=rocketNfreshReview+sellerrocketReview+importrocketReview+normalReview
    averageReview=int(totalReview/(normal+rocketNfresh+sellerrocket+importrocket))
    
    
    data={'keyword':input,'rocketNfresh':rocketNfresh,'sellerrocket':sellerrocket,'importrocket':importrocket,'normal':normal,'ad':ad,"totalReview":totalReview,
          "averageReview":averageReview,'rocketNfreshReview':rocketNfreshReview,'sellerrocketReview':sellerrocketReview,'importrocketReview':importrocketReview,
          "normalReview":normalReview,'minPrice':minPrice,'maxPrice':maxPrice,'diffPrice':diffPrice,'url':searchUrl}
    pprint.pprint(data)
    onlyvalue=list(data.values())
    pprint.pprint(onlyvalue)
    return data,onlyvalue,productInfoList
    
def createExcel():
    # Create a new workbook and select the active sheet
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Define the header titles and their respective spans
    headers = [
        ("순", 1),
        ("키워드", 1),
        ("상품수", 4),
        ("광고수", 1),
        ("리뷰수", 6),
        ("가격", 3),
        ("URL", 1)
    ]

    # Add headers to the first row and merge cells as required
    col_index = 1
    for title, span in headers:
        new_sheet.cell(row=1, column=col_index, value=title)
        if span > 1:
            new_sheet.merge_cells(start_row=1, end_row=1, start_column=col_index, end_column=col_index + span - 1)
        col_index += span

    # Define sub-headers for "상품수", "리뷰수", and "가격"
    sub_headers = [
        ("로켓배송(+프레쉬)", "판매자로켓", "로켓직구", "일반제품", ""),
        ("총리뷰수", "평균리뷰수", "로켓배송", "판매자로켓", "로켓직구", "일반제품"),
        ("최저가", "최고가", "가격차이")
    ]

    # Add sub-headers to the second row
    sub_col_index = 3  # Starting from the third column for "상품수"
    for sub_header_group in sub_headers:
        for sub_header in sub_header_group:
            new_sheet.cell(row=2, column=sub_col_index, value=sub_header)
            sub_col_index += 1

    # Adjust column width for readability
    for column in new_sheet.columns:
        max_length = 0
        column = list(column)  # Convert generator to list
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        new_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    # 모든 열의 너비를 15로 설정
    for column in new_sheet.columns:
        new_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 10

    # 1행과 2행의 모든 셀을 가운데 정렬
    for row in new_sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 셀 병합 (A1과 A2, B1과 B2, G1과 G2)
    new_sheet.merge_cells('A1:A2')
    new_sheet.merge_cells('B1:B2')
    new_sheet.merge_cells('G1:G2')
    new_sheet.merge_cells('Q1:Q2')

    # 노랑색 채우기 설정
    fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 1행과 2행을 노랑색으로 채우기
    for row in new_sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = fill_color

    # 테두리 스타일 설정
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )




    return new_sheet,new_wb



def GetGoogleSpreadSheet():
    scope = 'https://spreadsheets.google.com/feeds'
    json = 'source/credential.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json, scope)
    gc = gspread.authorize(credentials)
    sheet_url = 'https://docs.google.com/spreadsheets/d/1ESbtujllxTxsRSzN_4l6q2GW9j-Ygceghq_ZH8cyQ3I/edit#gid=0'
    doc = gc.open_by_url(sheet_url)
    worksheet = doc.worksheet('시트1')
    #=================전체정보가져오기
    all_data=worksheet.get_all_records()
    return all_data



def GetLogin(loginId,loginPw):
    result=False
    loginDatas=GetGoogleSpreadSheet()
    # print("loginDatas:",loginDatas,"/ loginDatas_TYPE:",type(loginDatas))
    macAdd=getmac.get_mac_address()
    print("macAdd:",macAdd,"/ macAdd_TYPE:",type(macAdd))
    timeNowTimestamp=datetime.datetime.now().timestamp()
    
    for loginData in loginDatas:
        startTimestamp = datetime.datetime.strptime(str(loginData['시작일']),"%Y%m%d").timestamp()
        endTimestamp=(datetime.datetime.strptime(str(loginData['종료일']),"%Y%m%d")+datetime.timedelta(days=1)).timestamp()
        if loginId==loginData['아이디'] and loginPw==loginData['비밀번호'] and startTimestamp<=timeNowTimestamp<=endTimestamp and macAdd==loginData['MAC ADDRESS']:
            print("로그인완료")
            result=True
            maxCount=loginData['상품갯수']
            print("maxCount:",maxCount,"/ maxCount_TYPE:",type(maxCount))
            break
        else:
            print("로그인실패")
            maxCount=0
    
    return result,maxCount


def FindKeyword(inputList,self,type):
    #==============키워드발굴
    # inputList=['오징어','닥스여성모자','왁','제주고사리','그란데캄포','전자다트핀']
    count=1
    columName=['순','키워드','상품수']
    new_sheet,new_wb=createExcel()
    timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    count=1
    productTotalList=[]
    for index,input in enumerate(inputList):
        text = "{}/{}번째 키워드 확인중...".format(index+1,len(inputList))
        if type=="TYPE1":
            print(text)
            self.user_signal.emit(text)
        else:
            print(text)
            self.user_signal2.emit(text)

        try:
            data,onlyvalue,productInfoList=GetSearch(input)
            result=[count]+onlyvalue
            new_sheet.append(result)
            new_wb.save('키워드발굴_{}.xlsx'.format(timeNow))
            count+=1
            text="완료"
            productTotalList.extend(productInfoList)
            with open('source/productTotalList.json', 'w',encoding='utf-8-sig') as f:
                json.dump(productTotalList, f, indent=2,ensure_ascii=False)
            if type=="TYPE1":
                print(text)
                self.user_signal.emit(text)
            else:
                print(text)
                self.user_signal2.emit(text)
        except:
            if type=="TYPE1":
                text="실패"
                print(text)
                self.user_signal.emit(text)
            else:
                text = "실패"
                print(text)
                self.user_signal2.emit(text)
        time.sleep(random.randint(10,20)*0.1)
        text = "============================="
        if type=="TYPE1":
            print(text)
            self.user_signal.emit(text)
        else:
            print(text)
            self.user_signal2.emit(text)
    # 테두리 스타일 설정
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # A열부터 P열까지의 모든 셀에 테두리 적용
    for row in new_sheet.iter_rows(min_col=1, max_col=17):  # A열(1)부터 P열(16)까지
        for cell in row:
            cell.border = thin_border

    # 1행과 2행의 모든 셀을 가운데 정렬
    for row in new_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print("하이퍼링크넣기")
    # 2행부터 끝행까지 반복
    for row in range(3, new_sheet.max_row + 1):
        # L열의 값을 읽음
        link_value = new_sheet[f'Q{row}'].value
        link_text = new_sheet[f'B{row}'].value

        # D열의 셀에 하이퍼링크 추가
        new_sheet[f'B{row}'].hyperlink = link_value
        new_sheet[f'B{row}'].value = link_text  # 표시될 텍스트
        new_sheet[f'B{row}'].font = Font(color='0000FF', underline='single')  # 링크처럼 보이게 스타일 설정

    new_wb.save('키워드발굴_{}.xlsx'.format(timeNow))

def FindBlueKeyword(filename,reviewCount,self,type):
    # reviewCount=100
    # filename='result_20231227_162323.xlsx'

    df=pd.read_excel(filename)
    # '열1'에서 nan 제거 후 리스트로 변환
    dataList = df['키워드'].dropna().tolist()
    with open ('source/productTotalList.json', "r",encoding='utf-8-sig') as f:
        productTotalList = json.load(f)
    # pprint.pprint(productTotalList)

    timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    wb=openpyxl.Workbook()
    new_sheet=wb.active
    columnName=['순','순위','키워드','상품명(링크)','가격','제품구분','도착일','분석최초리뷰일','총리뷰수','7일리뷰수','28일리뷰수','URL']
    new_sheet.append(columnName)
    
    count=1
    for index1,data in enumerate(dataList):
        print("data:",data)

        productInfos = [item for item in productTotalList if item['keyword'] == data]
        # productInfos=eval(data)
        for index2,productInfo in enumerate(productInfos):
            text="{}/{}키워드:{}/{}상품 리뷰 분석중...{}/{}".format(index1+1,len(dataList),index2+1,len(productInfos),productInfo['keyword'],productInfo['url'])
            print(text)
            self.user_signal2.emit(text)
            productId = productInfo['productNo']
            print("productId:",productId)
            with open('source/productInfo.json', 'w',encoding='utf-8-sig') as f:
                json.dump(productInfo, f, indent=2,ensure_ascii=False)
            try:
                productId=productInfo['productNo']
                reviewList,count7,count28,firstReview=GetReview(reviewCount,productId)
                data=[count,productInfo['rank'],productInfo['keyword'],productInfo['productName'],productInfo['price'],productInfo['category'],productInfo['arrivalInfo'],firstReview,productInfo['totalReview'],count7,count28,productInfo['url']]
                new_sheet.append(data)
                wb.save('블루키워드분석_{}.xlsx'.format(timeNow))
                count+=1
                text = "성공"
                print(text)
                self.user_signal2.emit(text)
            except:
                text = "실패"
                print(text)
                self.user_signal2.emit(text)
            text = "==================================="
            print(text)
            self.user_signal2.emit(text)
    # 테두리 스타일 설정
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # A열부터 P열까지의 모든 셀에 테두리 적용
    for row in new_sheet.iter_rows(min_col=1, max_col=12):  # A열(1)부터 P열(16)까지
        for cell in row:
            cell.border = thin_border

    # 1행과 2행의 모든 셀을 가운데 정렬
    for row in new_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 모든 열의 너비를 15로 설정
    for index,column in enumerate(new_sheet.columns):
        if index==3:
            new_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 50
        else:
            new_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 15

    # 노랑색 채우기 설정
    fill_color = PatternFill(start_color='008000', end_color='008000', fill_type='solid')

    # 1행과 2행을 노랑색으로 채우기
    for row in new_sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = fill_color

    # 2행부터 끝행까지 반복
    for row in range(2, new_sheet.max_row + 1):
        # L열의 값을 읽음
        link_value = new_sheet[f'L{row}'].value
        link_text = new_sheet[f'D{row}'].value

        # D열의 셀에 하이퍼링크 추가
        new_sheet[f'D{row}'].hyperlink = link_value
        new_sheet[f'D{row}'].value = link_text  # 표시될 텍스트
        new_sheet[f'D{row}'].font = Font(color='0000FF', underline='single')  # 링크처럼 보이게 스타일 설정

    wb.save('블루키워드분석_{}.xlsx'.format(timeNow))

def GetKeywords(filename):
    df=pd.read_excel(filename)
    dictData=df.to_dict(orient='records')
    pprint.pprint(dictData)
    return dictData



def GetCategory(value):
    resultList=[]
    cookies = {
        'NNB': 'VKDBZY7S2GHGK',
        'nx_ssl': '2',
        'ASID': 'dd8b11440000018cc47adfde00008e9a',
        '_ga': 'GA1.2.198916887.1704206835',
        'nid_inf': '934428385',
        'NID_AUT': 'ASWq16lkVaDwAIGLZ7/CfW1GsqCKNGKIEkl0WBRMEw3jRQV2aKreQDljHEUGdOd1',
        'NID_JKL': 'sJ0inAh371AkrW91gbC2FGVwde1C5iFD4zEhcmrEPgw=',
        'page_uid': 'iiGkRdqo1SCsseNrN38ssssssel-175807',
        'NID_SES': 'AAABkfqfJOxF2b23Y8F+9/NOvUiHqkwmn6q/P/QX81u/KC+Phma5fepavH47iSvwgEbgAAN2qdlqR0cCl7LMCLiP87n+KL4ovBEMYaatGVl8pUu2Xjv+tQJL/+qdHF6uh4w5vhAYpqvLMfatsW9xJ1ca0/2g//wtsJ6ga4vfj6Gr0XXx8+8rToGIg7GSXEsb4mnOe01esksbVf8OA12n/ZHW29GfAi4OVm96kweOf+Fv4OLBotuwxpXaS+p005miRSqJUJNBJp0NwXq6qhlMOp/zC6H4OkOCladKEhrzMrckzyNmKGaQcIxbj/Fj2D6ZRdYSCTMD46WOrBaAtixvTuFiDZ5WznB9M+crzqsS723e9h2yz2Nmz9pgXN6rre0NDC4N2h67ZcD5WMFGnbSB8G9iwbvMhCSfrtb6jOus6WiEYLPU3vAPb0AdBK39LIBRTe3d/9SwSbQ8qYBnbiNKOwEE8OBuckpf224tyWl5xhG5/Y0IJaTWFvuy4xfUA2sJ+t0PUpMcUh+2MjB43ZpO025JGwPf/lYBobhWtduOIcENkn2y',
        '_datalab_cid': '50000006',
    }
    headers = {
        'authority': 'datalab.naver.com',
        'accept': '*/*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'NNB=VKDBZY7S2GHGK; nx_ssl=2; ASID=dd8b11440000018cc47adfde00008e9a; _ga=GA1.2.198916887.1704206835; nid_inf=934428385; NID_AUT=ASWq16lkVaDwAIGLZ7/CfW1GsqCKNGKIEkl0WBRMEw3jRQV2aKreQDljHEUGdOd1; NID_JKL=sJ0inAh371AkrW91gbC2FGVwde1C5iFD4zEhcmrEPgw=; page_uid=iiGkRdqo1SCsseNrN38ssssssel-175807; NID_SES=AAABkfqfJOxF2b23Y8F+9/NOvUiHqkwmn6q/P/QX81u/KC+Phma5fepavH47iSvwgEbgAAN2qdlqR0cCl7LMCLiP87n+KL4ovBEMYaatGVl8pUu2Xjv+tQJL/+qdHF6uh4w5vhAYpqvLMfatsW9xJ1ca0/2g//wtsJ6ga4vfj6Gr0XXx8+8rToGIg7GSXEsb4mnOe01esksbVf8OA12n/ZHW29GfAi4OVm96kweOf+Fv4OLBotuwxpXaS+p005miRSqJUJNBJp0NwXq6qhlMOp/zC6H4OkOCladKEhrzMrckzyNmKGaQcIxbj/Fj2D6ZRdYSCTMD46WOrBaAtixvTuFiDZ5WznB9M+crzqsS723e9h2yz2Nmz9pgXN6rre0NDC4N2h67ZcD5WMFGnbSB8G9iwbvMhCSfrtb6jOus6WiEYLPU3vAPb0AdBK39LIBRTe3d/9SwSbQ8qYBnbiNKOwEE8OBuckpf224tyWl5xhG5/Y0IJaTWFvuy4xfUA2sJ+t0PUpMcUh+2MjB43ZpO025JGwPf/lYBobhWtduOIcENkn2y; _datalab_cid=50000006',
        'referer': 'https://datalab.naver.com/shoppingInsight/sCategory.naver',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }
    params = {
        'cid': value,
    }
    response = requests.get(
        'https://datalab.naver.com/shoppingInsight/getCategory.naver',
        params=params,
        cookies=cookies,
        headers=headers,
    )
    # print("response.text:",response.text,"/ response.text_TYPE:",type(response.text))
    results=json.loads(response.text)['childList']
    resultList.append({'name':"",'value':""})
    for result in results:
        value=result['cid']
        name=result['name']
        data={'name':name,'value':value}
        print("data:",data,"/ data_TYPE:",type(data))
        resultList.append(data)
    print("=================")
    return resultList

def GetTop500(targetCode,startDate,endDate,self):
    count=1
    resultList=[]
    wb = openpyxl.Workbook()
    ws = wb.active
    columName = ['keywords']
    ws.append(columName)
    timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    while True:

        cookies = {
            'NNB': 'VKDBZY7S2GHGK',
            'nx_ssl': '2',
            'ASID': 'dd8b11440000018cc47adfde00008e9a',
            '_ga': 'GA1.2.198916887.1704206835',
            'nid_inf': '934428385',
            'NID_AUT': 'ASWq16lkVaDwAIGLZ7/CfW1GsqCKNGKIEkl0WBRMEw3jRQV2aKreQDljHEUGdOd1',
            'NID_JKL': 'sJ0inAh371AkrW91gbC2FGVwde1C5iFD4zEhcmrEPgw=',
            'NID_SES': 'AAABlP78I4n+y04xCYkVXZdOvttgZsMF6D0gYMv25qW8ZINkvTj69m7BHKnWBAMv8UQRSV09d+2tZMsRxMIqdXP3DWkwM7Rgg9+0IJOpVN8/HcH6sErEokKL0sZsduhE13GVQgpI2aMJwJXoNXlaurk9bXEPcpK3nYWXTA8GCFDAytrlpwEwxT9TG50ZaxeMLQhowGKYSL3XinOrYFR4su+9YfzcX2BipXAJQ8r/bR9HGLWs2r/evwvNJOX8SXICbpfJx4bBg8p7IAHI+bousa2C6nKL8oNm+eB9eaYpw+LpLENCGBrhCeuep5ep+4uUUHNajbqdBV7+FRDetbg3jbCdJ1u6SSclPp40EtkiczSKxjf6vg3p4kBQm1grW5wZ1A8WTdHMxo8pfmGrZohpW3BmGxZPfYxGOhWZOJIReaQNNYMIIm18jSDZ8YthGpd7eJoxyMB9d1yRxHJk8m1jUZo2WW8yX9TRLRo3QNHol2RXzIbtshDDw8/QinhXmzPzEY+BkYUQnAd3I4+Kn13vpn0F176LCu+jCXouJTpdk96dwPwk',
            'page_uid': 'iiHc0wqo1awssbRCDM8ssssssaw-293305',
            '_datalab_cid': '50000009',
        }
    
        headers = {
            'authority': 'datalab.naver.com',
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            # 'cookie': 'NNB=VKDBZY7S2GHGK; nx_ssl=2; ASID=dd8b11440000018cc47adfde00008e9a; _ga=GA1.2.198916887.1704206835; nid_inf=934428385; NID_AUT=ASWq16lkVaDwAIGLZ7/CfW1GsqCKNGKIEkl0WBRMEw3jRQV2aKreQDljHEUGdOd1; NID_JKL=sJ0inAh371AkrW91gbC2FGVwde1C5iFD4zEhcmrEPgw=; NID_SES=AAABlP78I4n+y04xCYkVXZdOvttgZsMF6D0gYMv25qW8ZINkvTj69m7BHKnWBAMv8UQRSV09d+2tZMsRxMIqdXP3DWkwM7Rgg9+0IJOpVN8/HcH6sErEokKL0sZsduhE13GVQgpI2aMJwJXoNXlaurk9bXEPcpK3nYWXTA8GCFDAytrlpwEwxT9TG50ZaxeMLQhowGKYSL3XinOrYFR4su+9YfzcX2BipXAJQ8r/bR9HGLWs2r/evwvNJOX8SXICbpfJx4bBg8p7IAHI+bousa2C6nKL8oNm+eB9eaYpw+LpLENCGBrhCeuep5ep+4uUUHNajbqdBV7+FRDetbg3jbCdJ1u6SSclPp40EtkiczSKxjf6vg3p4kBQm1grW5wZ1A8WTdHMxo8pfmGrZohpW3BmGxZPfYxGOhWZOJIReaQNNYMIIm18jSDZ8YthGpd7eJoxyMB9d1yRxHJk8m1jUZo2WW8yX9TRLRo3QNHol2RXzIbtshDDw8/QinhXmzPzEY+BkYUQnAd3I4+Kn13vpn0F176LCu+jCXouJTpdk96dwPwk; page_uid=iiHc0wqo1awssbRCDM8ssssssaw-293305; _datalab_cid=50000009',
            'origin': 'https://datalab.naver.com',
            'referer': 'https://datalab.naver.com/shoppingInsight/sCategory.naver',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'x-requested-with': 'XMLHttpRequest',
        }
    
        data = {
            'cid': targetCode,
            'timeUnit': 'date',
            'startDate': startDate,
            'endDate': endDate,
            'age': '',
            'gender': '',
            'device': '',
            'page': str(count),
            'count': '20',
        }
        print("data:",data,"/ data_TYPE:",type(data))
    
        response = requests.post(
            'https://datalab.naver.com/shoppingInsight/getCategoryKeywordRank.naver',
            cookies=cookies,
            headers=headers,
            data=data,
        )
        print('statuscode:',response.status_code)
        
        results=json.loads(response.text)['ranks']
        pprint.pprint(results)
        if len(results)==0:
            break
        text = "{}-{}번 키워드 확인중...".format((count-1)*20,(count)*20)
        self.user_signal3.emit(text)
        print(text)
        for result in results:
            keyword=result['keyword']
            resultList.append(keyword)
            ws.append([keyword])
            wb.save('TOP500_{}.xlsx'.format(timeNow))
        with open('source/resultList.json', 'w',encoding='utf-8-sig') as f:
            json.dump(resultList, f, indent=2,ensure_ascii=False)
        count+=1
        time.sleep(random.randint(5,10)*0.1)


    



mymacadd=getmac.get_mac_address()
print("mymacadd:",mymacadd,"/ mymacadd_TYPE:",type(mymacadd))

class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal2 = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal3 = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,type,keywords,filename,maxCount,targetCode,startDate,endDate):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.type=type
        self.keywords=keywords
        self.filename=filename
        self.maxCount=maxCount
        self.targetCode=targetCode
        self.startDate=startDate
        self.endDate=endDate

    def run(self):
        if self.type=="TYPE1":
            FindKeyword(self.keywords,self,self.type)
        if self.type=="TYPE2":
            FindBlueKeyword(self.filename,self.maxCount,self,self.type)
        if self.type=="TYPE3":
            GetTop500(self.targetCode, self.startDate, self.endDate,self)
        text = "작업완료"
        self.user_signal.emit(text)

    def stop(self):
        pass




class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        with open ('source/idpw.json', "r",encoding='utf-8-sig') as f:
            idpw = json.load(f)
        self.lineEdit_4.setText(idpw['id'])
        self.lineEdit_5.setText(idpw['pw'])
        self.maxCount=30

        self.categoryList1 = [
            {'name':"",'value':""},
            {'name': '패션의류', 'value': '50000000'},
            {'name': '패션잡화', 'value': '50000001'},
            {'name': '화장품/미용', 'value': '50000002'},
            {'name': '디지털/가전', 'value': '50000003'},
            {'name': '가구/인테리어', 'value': '50000004'},
            {'name': '출산/육아', 'value': '50000005'},
            {'name': '식품', 'value': '50000006'},
            {'name': '스포츠/레저', 'value': '50000007'},
            {'name': '생활/건강', 'value': '50000008'},
            {'name': '여가/생활편의', 'value': '50000009'},
            {'name': '면세점', 'value': '50000010'},
            {'name': '도서', 'value': '50005542'}
        ]
        # ComboBox에 각 항목의 'name' 추가
        for item in self.categoryList1:
            self.comboBox.addItem(item['name'])

        # ComboBox 변경 이벤트 연결
        self.comboBox.currentIndexChanged.connect(self.on_combobox_changed)

        # ComboBox 변경 이벤트 연결
        self.comboBox_2.currentIndexChanged.connect(self.on_combobox_changed2)

        # ComboBox 변경 이벤트 연결
        self.comboBox_3.currentIndexChanged.connect(self.on_combobox_changed3)

        today = QDate.currentDate()
        self.dateEdit_2.setDate(today)

        # 현재 날짜 가져오기
        today = QDate.currentDate()

        # 한 달 전 날짜 계산
        one_month_ago = today.addMonths(-1)

        # 한 달 전 날짜 설정
        self.dateEdit.setDate(one_month_ago)

        self.categoryResult=[]

        self.startDate=""
        self.endDate=""

        self.keywords=""
        self.filename=""
        self.maxCount=99999

    def on_combobox_changed(self, index):
        # 선택된 항목의 value를 가져옴

        print(index)
        value = self.comboBox.currentText()

        for category1 in self.categoryList1:
            if value==category1['name']:
                input=category1['value']
                print("input:",input,"/ input_TYPE:",type(input))
                break

        # API 호출을 통해 데이터를 가져옴 (가상 함수)
        self.categoryList2 = GetCategory(input)


        # 두 번째 ComboBox 업데이트
        self.comboBox_2.clear()
        self.comboBox_3.clear()
        self.comboBox_4.clear()
        for item in self.categoryList2:
            self.comboBox_2.addItem(item['name'])



    def on_combobox_changed2(self, index):
        print("index:",index,"/ index_TYPE:",type(index))
        if index!=0:
            # 선택된 항목의 value를 가져옴

            value = self.comboBox_2.currentText()

            for category2 in self.categoryList2:
                if value == category2['name']:
                    input = category2['value']
                    print("input:", input, "/ input_TYPE:", type(input))
                    break

            # API 호출을 통해 데이터를 가져옴 (가상 함수)
            self.categoryList3 = GetCategory(input)

            # 두 번째 ComboBox 업데이트
            self.comboBox_3.clear()
            self.comboBox_4.clear()
            for item in self.categoryList3:
                self.comboBox_3.addItem(item['name'])

    def on_combobox_changed3(self, index):
        print("index:", index, "/ index_TYPE:", type(index))
        if index != 0:
            # 선택된 항목의 value를 가져옴

            value = self.comboBox_3.currentText()

            for category2 in self.categoryList3:
                if value == category2['name']:
                    input = category2['value']
                    print("input:", input, "/ input_TYPE:", type(input))
                    break

            # API 호출을 통해 데이터를 가져옴 (가상 함수)
            self.categoryList4 = GetCategory(input)

            # 두 번째 ComboBox 업데이트
            self.comboBox_4.clear()
            for item in self.categoryList4:
                self.comboBox_4.addItem(item['name'])



    def start1(self):
        rows = self.tableWidget.rowCount()
        values = []
        for row in range(rows):
            item = self.tableWidget.item(row, 0)
            if item is not None and item.text().strip():  # 빈칸이 아닌 경우에만 추가
                values.append(item.text().strip())
        print("values:",values,"/ values_TYPE:",type(values))
        # 중복 제거
        self.keywords=values
        self.keywords=self.keywords[:self.maxCount]

        print("self.keywords:",self.keywords,"/ self.keywords_TYPE:",type(self.keywords))
        self.filename=""
        self.maxCount=100
        self.type="TYPE1"
        self.x = Thread(self,self.type,self.keywords,self.filename,self.maxCount,self.categoryResult,self.startDate,self.endDate)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()
    def start2(self):
        self.type="TYPE2"
        self.keywords=[]
        self.filename=self.lineEdit_2.text()
        self.maxCount=int(self.lineEdit_3.text())
        self.x = Thread(self,self.type,self.keywords,self.filename,self.maxCount,self.categoryResult,self.startDate,self.endDate)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()
    def start3(self):
        self.type="TYPE3"
        date = self.dateEdit.date()
        self.startDate = date.toString("yyyy-MM-dd")
        date = self.dateEdit_2.date()
        self.endDate = date.toString("yyyy-MM-dd")

        self.category1Name=self.comboBox.currentText()
        if len(self.category1Name)>=1:
            self.categoryResult.append(self.category1Name)
        self.category2Name = self.comboBox_2.currentText()
        if len(self.category2Name)>=1:
            self.categoryResult.append(self.category2Name)
        self.category3Name = self.comboBox_3.currentText()
        if len(self.category3Name)>=1:
            self.categoryResult.append(self.category3Name)
        self.category4Name = self.comboBox_4.currentText()
        if len(self.category4Name)>=1:
            self.categoryResult.append(self.category4Name)
        
        self.categoryLength=len(self.categoryResult)
        
        self.targetCategory=self.categoryResult[-1]
        if self.categoryLength==1:
            for data in self.categoryList1:
                if data['name']==self.targetCategory:
                    targetCode=data['value']
                    break
        elif self.categoryLength==2:
            for data in self.categoryList2:
                if data['name']==self.targetCategory:
                    targetCode=data['value']
                    break
        elif self.categoryLength==3:
            for data in self.categoryList3:
                if data['name']==self.targetCategory:
                    targetCode=data['value']
                    break
        elif self.categoryLength==4:
            for data in self.categoryList4:
                if data['name']==self.targetCategory:
                    targetCode=data['value']
                    break
        print("targetCode:",targetCode,"/ targetCode_TYPE:",type(targetCode))

        print("self.categoryResult:",self.categoryResult,"/ self.categoryResult_TYPE:",type(self.categoryResult))

        self.x = Thread(self,self.type,self.keywords,self.filename,self.maxCount,targetCode,self.startDate,self.endDate)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal3.connect(self.slot3)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()
    def find1(self):
        print("find")
        self.fname = QFileDialog.getOpenFileName(self, "Open file", './')[0]
        print(self.fname)
        self.lineEdit.setText(self.fname)

        try:
            keywords=GetKeywords(self.fname)

            self.tableWidget.setRowCount(1000)
            self.tableWidget.setColumnCount(1)
            self.tableWidget.setColumnWidth(0, 800)
            self.tableWidget.setHorizontalHeaderLabels(['키워드'])
            # 키워드를 QTableWidget에 추가
            for i, keyword in enumerate(keywords):
                self.tableWidget.setItem(i, 0, QTableWidgetItem(keyword['keywords']))
        except:
            print("파일없음")

    def find2(self):
        print("find")
        self.filename = QFileDialog.getOpenFileName(self, "Open file", './')[0]
        self.lineEdit_2.setText(self.filename)
    def login(self):
        loginId=self.lineEdit_4.text()
        loginPw=self.lineEdit_5.text()
        print("loginId:",loginId,"/ loginId_TYPE:",type(loginId))
        print("loginPw:",loginPw,"/ loginPw_TYPE:",type(loginPw))
        result,self.maxCount=GetLogin(loginId,loginPw)

        print("result:",result,"/ result_TYPE:",type(result))

        if result==True:
            QMessageBox.information(self, "완료창", "로그인 완료")
            self.pushButton_3.setEnabled(True)
            self.pushButton_4.setEnabled(True)
            self.pushButton_5.setEnabled(True)
            self.pushButton_6.setEnabled(True)
            self.pushButton_14.setEnabled(True)
        else:
            QMessageBox.information(self, "에러창", "로그인 실패")

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))
    def slot2(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit_2.append(str(data1))
    def slot3(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit_5.append(str(data1))

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())




